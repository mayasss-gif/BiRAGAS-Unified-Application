"""
ExcelExporter — Multi-Tab Excel Export for BiRAGAS CRISPR Complete
====================================================================
Exports ALL analysis results to professional Excel with color-coded tabs.
"""

import logging
import os
from typing import Any, Dict, List, Optional

logger = logging.getLogger("biragas_crispr.reporting.excel")

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    OPENPYXL_OK = True
except ImportError:
    OPENPYXL_OK = False


class ExcelExporter:
    """Export all BiRAGAS CRISPR results to multi-tab Excel."""

    NAVY = "1A237E"
    BLUE = "E3F2FD"
    PURPLE = "F3E5F5"
    GREEN = "E8F5E9"
    ORANGE = "FFF3E0"
    RED = "FCE4EC"

    def __init__(self):
        if not OPENPYXL_OK:
            raise ImportError("pip install openpyxl")

    def export(self, report_data: Dict, output_path: str, disease: str = "Disease") -> str:
        """Export complete analysis to Excel."""
        wb = openpyxl.Workbook()

        # Tab 1: Summary
        self._tab_summary(wb.active, report_data, disease)

        # Tab 2: Knockout Rankings
        ko = report_data.get('knockout', report_data.get('dna_stages', {}).get('knockout', {}))
        if ko:
            self._tab_knockouts(wb.create_sheet(), ko)

        # Tab 3: Combinations
        combos = report_data.get('combinations', {})
        if combos:
            self._tab_combinations(wb.create_sheet(), combos)

        # Tab 4: Guide Design
        guides = report_data.get('guides', {})
        if guides:
            self._tab_guides(wb.create_sheet(), guides)

        # Tab 5: Causality
        caus = report_data.get('causality', {})
        if caus:
            self._tab_causality(wb.create_sheet(), caus)

        # Tab 6: Scale
        scale = report_data.get('scale', {})
        if scale:
            self._tab_scale(wb.create_sheet(), scale)

        os.makedirs(os.path.dirname(output_path) or '.', exist_ok=True)
        wb.save(output_path)
        logger.info(f"Excel exported: {output_path} ({len(wb.sheetnames)} tabs)")
        return output_path

    def _setup_sheet(self, ws, title, headers, widths):
        ws.title = title
        hdr_font = Font(color="FFFFFF", bold=True, size=10)
        hdr_fill = PatternFill(start_color=self.NAVY, end_color=self.NAVY, fill_type="solid")
        for i, h in enumerate(headers, 1):
            c = ws.cell(1, i, h)
            c.font = hdr_font
            c.fill = hdr_fill
            c.alignment = Alignment(horizontal='center', wrap_text=True)
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w
        ws.freeze_panes = "A2"

    def _tab_summary(self, ws, data, disease):
        self._setup_sheet(ws, "Summary", ["Metric", "Value"], [30, 50])
        rows = [
            ["Disease", disease],
            ["Pipeline", data.get('pipeline', 'BiRAGAS CRISPR Complete v3.0')],
            ["Duration (s)", data.get('duration_seconds', 'N/A')],
            ["Genes", data.get('scale', {}).get('genes', 'N/A')],
            ["DNA Configs", data.get('scale', {}).get('dna_configs', 'N/A')],
            ["RNA Configs", data.get('scale', {}).get('rna_configs', 'N/A')],
            ["Total Configs", data.get('scale', {}).get('total_configs', 'N/A')],
            ["Total Combinations", data.get('scale', {}).get('total_combos', 'N/A')],
            ["Causality Modules Run", data.get('causality', {}).get('modules_run', 'N/A')],
            ["Causality Failures", data.get('causality', {}).get('modules_failed', 'N/A')],
            ["Errors", len(data.get('errors', []))],
        ]
        for i, (metric, val) in enumerate(rows, 2):
            ws.cell(i, 1, metric).font = Font(bold=True, size=10)
            ws.cell(i, 2, str(val)).font = Font(size=10)

    def _tab_knockouts(self, ws, ko):
        top = ko.get('top_15', ko.get('top_5', []))
        self._setup_sheet(ws, "Knockouts", ["Rank", "Gene", "Ensemble", "Direction", "Confidence", "CI Low", "CI High"], [8, 14, 12, 14, 12, 10, 10])
        for i, t in enumerate(top, 2):
            ws.cell(i, 1, i - 1)
            ws.cell(i, 2, t.get('gene', ''))
            ws.cell(i, 3, round(t.get('ensemble', 0), 4))
            ws.cell(i, 4, t.get('direction', ''))
            ws.cell(i, 5, round(t.get('confidence', 0), 3))
            ci = t.get('ci', [0, 0])
            ws.cell(i, 6, round(ci[0], 4) if len(ci) > 0 else 0)
            ws.cell(i, 7, round(ci[1], 4) if len(ci) > 1 else 0)

    def _tab_combinations(self, ws, combos):
        self._setup_sheet(ws, "Combinations", ["Class", "Gene A", "Gene B", "Synergy", "Type", "Models", "Cross Bonus"], [10, 12, 12, 10, 12, 8, 12])
        row = 2
        for cls_key in ['dna_x_dna', 'rna_x_rna', 'dna_x_rna']:
            top = combos.get(cls_key, {}).get('top_5', [])
            cls_name = cls_key.replace('_x_', '×').upper()
            for c in top:
                ws.cell(row, 1, cls_name)
                ws.cell(row, 2, c['genes'][0] if len(c['genes']) > 0 else '')
                ws.cell(row, 3, c['genes'][1] if len(c['genes']) > 1 else '')
                ws.cell(row, 4, round(c.get('synergy', 0), 4))
                ws.cell(row, 5, c.get('type', ''))
                ws.cell(row, 6, c.get('models_used', 0))
                ws.cell(row, 7, round(c.get('cross_modal_bonus', 0), 3))
                row += 1

    def _tab_guides(self, ws, guides):
        self._setup_sheet(ws, "Guides", ["Gene", "DNA Configs", "Max KO%", "RNA Configs", "Max KD%"], [14, 12, 10, 12, 10])
        for i, (gene, gd) in enumerate(guides.items(), 2):
            ws.cell(i, 1, gene)
            ws.cell(i, 2, gd.get('dna', {}).get('configs', 0))
            ws.cell(i, 3, round(gd.get('dna', {}).get('max_ko', 0) * 100, 1))
            ws.cell(i, 4, gd.get('rna', {}).get('configs', 0))
            ws.cell(i, 5, round(gd.get('rna', {}).get('max_kd', 0) * 100, 1))

    def _tab_causality(self, ws, caus):
        self._setup_sheet(ws, "Causality", ["Phase", "Modules Run", "Failed", "Status"], [20, 14, 10, 12])
        phases = caus.get('phases', {})
        for i, (pk, pd) in enumerate(phases.items(), 2):
            ws.cell(i, 1, pk)
            ws.cell(i, 2, pd.get('modules_run', 0))
            ws.cell(i, 3, pd.get('modules_failed', 0))
            ws.cell(i, 4, "PASS" if pd.get('modules_failed', 0) == 0 else "FAIL")
            if pd.get('modules_failed', 0) > 0:
                ws.cell(i, 4).font = Font(color="D32F2F", bold=True)

    def _tab_scale(self, ws, scale):
        self._setup_sheet(ws, "Scale", ["Metric", "Value"], [25, 25])
        for i, (k, v) in enumerate(scale.items(), 2):
            ws.cell(i, 1, k.replace('_', ' ').title())
            ws.cell(i, 2, str(v))
